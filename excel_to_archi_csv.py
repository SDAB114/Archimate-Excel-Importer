"""
Excel to Archi CSV Importer
============================
Leest een Excel-bestand met twee kolommen en genereert
een CSV die je in Archi kunt importeren via:
  File > Import > CSV

Verwacht Excel-formaat:
  Kolom A: Applicatie     — naam van de applicatie
  Kolom B: Grouping-Node  — naam van de grouping waartoe de applicatie behoort

Het script maakt:
- Grouping-elementen aan (uniek, geen duplicaten)
- Application Components aan voor elke applicatie
- Composition-relaties van elke Grouping naar de bijbehorende applicaties

Gebruik:
  python excel_to_archi_csv.py

Vereisten:
  pip install openpyxl
"""

import csv
import uuid
import os
import openpyxl

# ─────────────────────────────────────────────
# CONFIGURATIE — pas deze waarden aan
# ─────────────────────────────────────────────

EXCEL_BESTAND     = "c:\Archi\Import\Applicatie-import.xlsx"  # Pad naar je Excel-bestand
KOLOM_APPLICATIE  = "Applicatie"        # Kolomnaam kolom A
KOLOM_GROUPING    = "Grouping-Node"     # Kolomnaam kolom B
SHEET_NAAM        = None                # None = eerste sheet; of bijv. "Sheet1"

# Uitvoermap voor de CSV-bestanden (Archi importeert een map met 3 bestanden)
UITVOER_MAP       = "c:\Archi\Import"

# ─────────────────────────────────────────────
# HULPFUNCTIES
# ─────────────────────────────────────────────

def nieuw_id():
    """Genereert een uniek Archi-compatibel ID."""
    return "id-" + str(uuid.uuid4())


def lees_excel(excel_bestand, kolom_applicatie, kolom_grouping, sheet_naam=None):
    """
    Leest applicatienamen en bijbehorende grouping-nodes uit Excel.
    Geeft een lijst van tuples terug: [(applicatie, grouping), ...]
    """
    wb = openpyxl.load_workbook(excel_bestand)
    ws = wb[sheet_naam] if sheet_naam else wb.active

    # Zoek kolomindices op basis van kolomnamen in rij 1
    headers = [cell.value for cell in ws[1]]

    for kolom in [kolom_applicatie, kolom_grouping]:
        if kolom not in headers:
            raise ValueError(
                f"Kolom '{kolom}' niet gevonden. "
                f"Beschikbare kolommen: {headers}"
            )

    idx_app      = headers.index(kolom_applicatie)
    idx_grouping = headers.index(kolom_grouping)

    rijen = []
    for rij in ws.iter_rows(min_row=2, values_only=True):
        applicatie = rij[idx_app]
        grouping   = rij[idx_grouping]
        if applicatie:
            rijen.append((
                str(applicatie).strip(),
                str(grouping).strip() if grouping else "Overig"
            ))

    print(f"✓ {len(rijen)} applicaties gevonden in '{excel_bestand}'")
    return rijen


def genereer_csv(rijen, uitvoer_map):
    """
    Genereert de drie CSV-bestanden die Archi verwacht:
      - elements.csv   — Grouping-nodes + Application Components
      - relations.csv  — Composition-relaties
      - properties.csv — leeg (uitbreidbaar)
    """
    os.makedirs(uitvoer_map, exist_ok=True)

    # ── 1. Verwerk unieke Grouping-nodes ────────────────────────
    # Gebruik een dict zodat elke Grouping maar één keer voorkomt
    grouping_map = {}  # { grouping_naam: grouping_id }
    for _, grouping_naam in rijen:
        if grouping_naam not in grouping_map:
            grouping_map[grouping_naam] = nieuw_id()

    print(f"✓ {len(grouping_map)} unieke Grouping-nodes gevonden: "
          f"{', '.join(grouping_map.keys())}")

    # ── 2. Elements CSV ──────────────────────────────────────────
    elementen = []

    # Grouping-elementen
    for naam, gid in grouping_map.items():
        elementen.append({
            "ID":            gid,
            "Type":          "Grouping",
            "Name":          naam,
            "Documentation": ""
        })

    # Application Components
    app_elementen = []
    for applicatie_naam, grouping_naam in rijen:
        app_elementen.append({
            "ID":            nieuw_id(),
            "Type":          "ApplicationComponent",
            "Name":          applicatie_naam,
            "Documentation": "",
            "_grouping_id":  grouping_map[grouping_naam]  # intern, niet in CSV
        })
    elementen.extend(app_elementen)

    elements_pad = os.path.join(uitvoer_map, "elements.csv")
    with open(elements_pad, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["ID", "Type", "Name", "Documentation"],
            extrasaction="ignore"  # negeert het interne _grouping_id veld
        )
        writer.writeheader()
        writer.writerows(elementen)
    print(f"✓ elements.csv aangemaakt ({len(elementen)} elementen)")

    # ── 3. Relations CSV ─────────────────────────────────────────
    relaties = []
    for app in app_elementen:
        relaties.append({
            "ID":            nieuw_id(),
            "Type":          "CompositionRelationship",
            "Name":          "",
            "Documentation": "",
            "Source":        app["_grouping_id"],
            "Target":        app["ID"]
        })

    relations_pad = os.path.join(uitvoer_map, "relations.csv")
    with open(relations_pad, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["ID", "Type", "Name", "Documentation", "Source", "Target"]
        )
        writer.writeheader()
        writer.writerows(relaties)
    print(f"✓ relations.csv aangemaakt ({len(relaties)} relaties)")

    # ── 4. Properties CSV ────────────────────────────────────────
    properties_pad = os.path.join(uitvoer_map, "properties.csv")
    with open(properties_pad, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["ID", "Key", "Value"])
        writer.writeheader()
    print(f"✓ properties.csv aangemaakt (leeg)")

    return uitvoer_map


# ─────────────────────────────────────────────
# HOOFDPROGRAMMA
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 50)
    print("  Excel → Archi CSV Importer")
    print("=" * 50)

    # Stap 1: Lees Excel
    rijen = lees_excel(EXCEL_BESTAND, KOLOM_APPLICATIE, KOLOM_GROUPING, SHEET_NAAM)

    # Stap 2: Genereer CSV
    uitvoer = genereer_csv(rijen, UITVOER_MAP)

    print()
    print("=" * 50)
    print(f"✅ Klaar! CSV-bestanden staan in: ./{uitvoer}/")
    print()
    print("Importeer in Archi via:")
    print("  File > Import > CSV")
    print(f"  Selecteer de map: {uitvoer}/")
    print("=" * 50)
