"""
Excel to Archi XML Importer
============================
Leest een Excel-bestand met applicaties en grouping-nodes en voegt
de applicaties toe aan een bestaand Archi .archimate model.

Verwacht Excel-formaat:
  Kolom A: Applicatie     — naam van de applicatie
  Kolom B: Grouping-Node  — naam van de grouping waartoe de applicatie behoort

Gebruik:
  python excel_to_archi_xml.py

Vereisten:
  pip install openpyxl lxml
"""

import uuid
import os
import zipfile
import io
import openpyxl
from lxml import etree

# ─────────────────────────────────────────────
# CONFIGURATIE — pas deze waarden aan
# ─────────────────────────────────────────────

EXCEL_BESTAND     = r"C:\Archi\Import\Applicatie-import.xlsx"
ARCHIMATE_BESTAND = r"C:\Archi\Bravis.archimate"
UITVOER_BESTAND   = r"C:\Archi\Bravis_updated.archimate"

KOLOM_APPLICATIE  = "Applicatie"
KOLOM_GROUPING    = "Grouping-Node"
SHEET_NAAM        = None   # None = eerste sheet

VIEW_NAAM         = "Applicatiefunctiemodel"

# Afmetingen van een Application Component blokje in de view
APP_BREEDTE   = 120
APP_HOOGTE    = 55
APP_MARGE     = 10

# Namespace
XSI = "http://www.w3.org/2001/XMLSchema-instance"
NS  = "http://www.archimatetool.com/archimate"

# ─────────────────────────────────────────────
# HULPFUNCTIES
# ─────────────────────────────────────────────

def nieuw_id():
    return "id-" + str(uuid.uuid4())


def lees_excel(excel_bestand, kolom_applicatie, kolom_grouping, sheet_naam=None):
    wb = openpyxl.load_workbook(excel_bestand)
    ws = wb[sheet_naam] if sheet_naam else wb.active
    headers = [cell.value for cell in ws[1]]
    for kolom in [kolom_applicatie, kolom_grouping]:
        if kolom not in headers:
            raise ValueError(f"Kolom '{kolom}' niet gevonden. Beschikbare kolommen: {headers}")
    idx_app      = headers.index(kolom_applicatie)
    idx_grouping = headers.index(kolom_grouping)
    rijen = []
    for rij in ws.iter_rows(min_row=2, values_only=True):
        applicatie = rij[idx_app]
        grouping   = rij[idx_grouping]
        if applicatie:
            rijen.append((
                str(applicatie).strip(),
                str(grouping).strip() if grouping else "Overig"
            ))
    print(f"✓ {len(rijen)} applicaties gelezen uit Excel")
    return rijen


def lees_archimate(archimate_bestand):
    parser = etree.XMLParser(remove_blank_text=True)
    if zipfile.is_zipfile(archimate_bestand):
        print("✓ Gecomprimeerd .archimate bestand gedetecteerd — uitpakken...")
        with zipfile.ZipFile(archimate_bestand, "r") as z:
            xml_naam = [f for f in z.namelist() if f.endswith(".xml")][0]
            with z.open(xml_naam) as f:
                inhoud = f.read()
        tree = etree.parse(io.BytesIO(inhoud), parser)
    else:
        tree = etree.parse(archimate_bestand, parser)
    root = tree.getroot()
    print(f"✓ Model geladen")
    return tree, root


def zoek_folder(root, folder_type):
    """Zoekt een folder op type (bijv. 'application', 'relations')."""
    for el in root:
        if el.tag == "folder" and el.get("type") == folder_type:
            return el
    return None


def zoek_elementen_folder(root):
    """Zoekt de 'Elementen' subfolder binnen de Application folder."""
    app_folder = zoek_folder(root, "application")
    if app_folder is None:
        return root
    for child in app_folder:
        if child.tag == "folder" and child.get("name") == "Elementen":
            return child
    return app_folder


def zoek_relaties_folder(root):
    """Zoekt de Relations folder."""
    rel_folder = zoek_folder(root, "relations")
    return rel_folder if rel_folder is not None else root


def zoek_element_op_naam(root, naam, xsi_type=None):
    """Zoekt een element op naam (case-insensitief) en optioneel xsi:type."""
    for el in root.iter():
        el_naam = el.get("name") or ""
        if el_naam.lower() != naam.lower():
            continue
        if xsi_type:
            el_type = el.get(f"{{{XSI}}}type") or ""
            if xsi_type.lower() not in el_type.lower():
                continue
        return el
    return None


def voeg_app_toe_aan_model(root, naam):
    """Voegt een ApplicationComponent toe aan de Elementen folder. Geeft id terug."""
    # Controleer of het al bestaat
    bestaand = zoek_element_op_naam(root, naam, "ApplicationComponent")
    if bestaand is not None:
        print(f"  ~ Al aanwezig in model: {naam}")
        return bestaand.get("id")

    folder = zoek_elementen_folder(root)
    nieuw_el = etree.SubElement(folder, "element")
    nieuw_el.set(f"{{{XSI}}}type", "archimate:ApplicationComponent")
    nieuw_el.set("name", naam)
    eid = nieuw_id()
    nieuw_el.set("id", eid)
    print(f"  + Element aangemaakt: {naam}")
    return eid


def voeg_relatie_toe_aan_model(root, source_id, target_id):
    """Voegt een CompositionRelationship toe. Geeft id terug."""
    # Controleer of het al bestaat
    for el in root.iter():
        el_type = el.get(f"{{{XSI}}}type") or ""
        if ("CompositionRelationship" in el_type and
                el.get("source") == source_id and
                el.get("target") == target_id):
            return el.get("id")

    folder = zoek_relaties_folder(root)
    rel = etree.SubElement(folder, "element")
    rel.set(f"{{{XSI}}}type", "archimate:CompositionRelationship")
    rid = nieuw_id()
    rel.set("id",     rid)
    rel.set("source", source_id)
    rel.set("target", target_id)
    print(f"  + Relatie aangemaakt")
    return rid


def zoek_view(root, view_naam):
    """Zoekt de view op naam."""
    for el in root.iter():
        if el.get("name") == view_naam:
            print(f"✓ View gevonden: '{view_naam}'")
            return el
    raise ValueError(f"View '{view_naam}' niet gevonden in het model!")


def zoek_grouping_in_view(view, grouping_id):
    """Zoekt het view-child element dat verwijst naar de grouping via archimateElement."""
    for el in view.iter():
        ref = el.get("archimateElement") or el.get("elementRef")
        if ref == grouping_id:
            return el
    return None


def voeg_app_toe_aan_view(grouping_view_el, app_id, positie_index):
    """Plaatst een Application Component visueel binnen het Grouping-blok."""
    # Controleer of al aanwezig
    for el in grouping_view_el.iter():
        if el.get("archimateElement") == app_id:
            print(f"  ~ Al aanwezig in view")
            return

    # Haal bounds van de grouping op
    bounds = grouping_view_el.find("bounds")
    if bounds is not None:
        gx = int(bounds.get("x", 0))
        gy = int(bounds.get("y", 0))
    else:
        gx, gy = 0, 0

    # Bereken positie (4 per rij)
    kolom  = positie_index % 4
    rij_nr = positie_index // 4
    x = gx + APP_MARGE + kolom * (APP_BREEDTE + APP_MARGE)
    y = gy + 30 + rij_nr * (APP_HOOGTE + APP_MARGE)

    child = etree.SubElement(grouping_view_el, "child")
    child.set(f"{{{XSI}}}type", "archimate:DiagramObject")
    child.set("id",               nieuw_id())
    child.set("archimateElement", app_id)

    b = etree.SubElement(child, "bounds")
    b.set("x",      str(x))
    b.set("y",      str(y))
    b.set("width",  str(APP_BREEDTE))
    b.set("height", str(APP_HOOGTE))

    print(f"  → Geplaatst in view (kolom {kolom+1}, rij {rij_nr+1})")


def sla_op_als_zip(tree, uitvoer_bestand):
    """Slaat het bijgewerkte model op als gecomprimeerd .archimate bestand."""
    xml_bytes = etree.tostring(tree, pretty_print=True, xml_declaration=True, encoding="UTF-8")
    with zipfile.ZipFile(uitvoer_bestand, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("model.xml", xml_bytes)
    print(f"✓ Opgeslagen als: {uitvoer_bestand}")


# ─────────────────────────────────────────────
# HOOFDPROGRAMMA
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  Excel → Archi XML Importer")
    print("=" * 55)

    # 1. Lees Excel
    rijen = lees_excel(EXCEL_BESTAND, KOLOM_APPLICATIE, KOLOM_GROUPING, SHEET_NAAM)

    # 2. Laad model
    tree, root = lees_archimate(ARCHIMATE_BESTAND)

    # 3. Zoek view
    view = zoek_view(root, VIEW_NAAM)

    # 4. Verwerk elke rij
    teller_per_grouping = {}

    for applicatie_naam, grouping_naam in rijen:
        print(f"\nVerwerk: {applicatie_naam} → {grouping_naam}")

        # Zoek grouping in model
        grouping_el = zoek_element_op_naam(root, grouping_naam)
        if grouping_el is None:
            print(f"  ⚠ Grouping '{grouping_naam}' niet gevonden in model — overgeslagen")
            continue
        grouping_id = grouping_el.get("id")
        print(f"  ✓ Grouping gevonden: '{grouping_el.get('name')}'")

        # Voeg ApplicationComponent toe aan model
        app_id = voeg_app_toe_aan_model(root, applicatie_naam)

        # Voeg Composition-relatie toe
        voeg_relatie_toe_aan_model(root, grouping_id, app_id)

        # Zoek grouping in view
        grouping_view_el = zoek_grouping_in_view(view, grouping_id)
        if grouping_view_el is None:
            print(f"  ⚠ Grouping niet gevonden in view — overgeslagen")
            continue

        # Positie binnen grouping
        idx = teller_per_grouping.get(grouping_naam, 0)
        teller_per_grouping[grouping_naam] = idx + 1

        # Voeg toe aan view
        voeg_app_toe_aan_view(grouping_view_el, app_id, idx)

    # 5. Sla op als zip (zoals Archi verwacht)
    sla_op_als_zip(tree, UITVOER_BESTAND)

    print()
    print("=" * 55)
    print(f"✅ Klaar!")
    print(f"Open in Archi: File → Open → {UITVOER_BESTAND}")
    print("=" * 55)
